"""成本核算系统 — CSV/Excel 导出模块

从 cost_accounting.py 拆分，减少主文件行数（980→≤600行）。
"""
import logging
from io import StringIO, BytesIO
from datetime import datetime
from typing import Optional, Any

from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session, joinedload
from app.core.database import get_db
from app.core.security import require_menu
from app.models.cost_accounting import CostAccountingSheet, CostAccountingItem

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/cost-accounting", tags=["cost-accounting-export"])
_DEP = Depends(require_menu("cost-accounting"))


@router.get("/reports/export/csv")
def export_csv(
    sheet_id: int = Query(...),
    db: Session = Depends(get_db),
    _=_DEP,
):
    """导出核算单CSV"""
    s = db.query(CostAccountingSheet).filter(CostAccountingSheet.id == sheet_id).first()
    if not s:
        raise HTTPException(404, "核算单不存在")
    items = db.query(CostAccountingItem).filter(CostAccountingItem.sheet_id == s.id).all()
    import csv
    output = StringIO()
    w = csv.writer(output)
    w.writerow(["核算单", s.sheet_no, "", "", "", ""])
    w.writerow(["成本类别", "项目名称", "目标金额", "实际金额", "差异", "差异率%"])
    w.writerow(["物料成本", "BOM物料合计", s.material_cost_target, s.material_cost_actual,
                 s.material_cost_actual - s.material_cost_target, ""])
    for i in items:
        w.writerow([i.cost_category, i.item_name, i.target_amount, i.actual_amount, i.variance, i.variance_pct])
    w.writerow(["合计", "", s.total_cost_target, s.total_cost_actual, s.variance_amount, s.variance_pct])
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={s.sheet_no}.csv"},
    )


@router.get("/reports/export/excel")
def export_excel(
    period_id: Optional[int] = Query(None, description="筛选期间ID"),
    status: Optional[str] = Query(None, description="筛选状态: draft/finalized"),
    sheet_ids: Optional[str] = Query(None, description="指定核算单ID列表，逗号分隔"),
    plan_id: Optional[str] = Query(None, description="筛选产品策划ID"),
    db: Session = Depends(get_db),
    _=_DEP,
) -> Any:
    """批量导出核算单为Excel（.xlsx）

    支持按期间/状态筛选，或按ID列表导出。
    每个核算单为Excel中的一个sheet，含明细项。
    自动按差异率排序，差异率为负时标红。
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse

    # ── 1. 查询核算单 ──
    from sqlalchemy.orm import joinedload as jl
    query = db.query(CostAccountingSheet).options(
        jl(CostAccountingSheet.product_plan),
        jl(CostAccountingSheet.period),
    ).order_by(CostAccountingSheet.period_id, CostAccountingSheet.id)

    if sheet_ids:
        ids = [int(x.strip()) for x in sheet_ids.split(",") if x.strip().isdigit()]
        if not ids:
            raise HTTPException(400, "sheet_ids 参数格式无效，请提供逗号分隔的数字ID列表")
        query = query.filter(CostAccountingSheet.id.in_(ids))
    else:
        if period_id:
            query = query.filter(CostAccountingSheet.period_id == period_id)
        if status:
            query = query.filter(CostAccountingSheet.status == status)
        if plan_id:
            query = query.filter(CostAccountingSheet.product_plan_id == plan_id)

    sheets = query.all()
    if not sheets:
        raise HTTPException(404, "没有符合条件的核算单")

    # ── 2. 批量获取明细 ──
    sheet_ids_list = [s.id for s in sheets]
    items_map: dict[int, list[CostAccountingItem]] = {sid: [] for sid in sheet_ids_list}
    all_items = (
        db.query(CostAccountingItem)
        .filter(CostAccountingItem.sheet_id.in_(sheet_ids_list))
        .order_by(CostAccountingItem.sheet_id, CostAccountingItem.cost_category)
        .all()
    )
    for item in all_items:
        items_map.setdefault(item.sheet_id, []).append(item)

    # ── 3. 构建 Excel ──
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 样式定义
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_align_right = Alignment(horizontal="right", vertical="center")
    data_align_center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    title_font = Font(name="微软雅黑", bold=True, size=14, color="1F4E79")
    subtitle_font = Font(name="微软雅黑", size=10, color="595959")
    red_font = Font(name="微软雅黑", color="FF0000")
    total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    total_font = Font(name="微软雅黑", bold=True, size=11)
    cat_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    categories_cn = {"material": "物料成本", "labor": "人工成本", "overhead": "制造费用"}

    for idx, sheet in enumerate(sheets):
        plan_name = sheet.product_plan.name if sheet.product_plan else ""
        sheet_name = f"{sheet.sheet_no[-8:]}" if len(sheet.sheet_no) > 8 else f"Sheet{idx+1}"
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]

        ws = wb.create_sheet(title=sheet_name)

        # 标题行
        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        title_cell.value = f"成本核算单 — {sheet.sheet_no}"
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 30

        # 副标题
        ws.merge_cells("A2:F2")
        sub = ws["A2"]
        period_name = sheet.period.period_name if sheet.period else f"期间#{sheet.period_id}"
        sub.value = f"{plan_name}  |  期间: {period_name}  |  状态: {'已定稿' if sheet.status == 'finalized' else '草稿'}  |  导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        sub.font = subtitle_font
        ws.row_dimensions[2].height = 22

        # 汇总表头
        headers = ["成本类别", "项目名称", "目标金额(元)", "实际金额(元)", "差异(元)", "差异率(%)"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        ws.row_dimensions[4].height = 24

        # 成本概要行
        row = 5
        summary_data = [
            ("物料成本", "BOM物料合计", sheet.material_cost_target, sheet.material_cost_actual,
             sheet.material_cost_actual - sheet.material_cost_target,
             round((sheet.material_cost_actual - sheet.material_cost_target) / sheet.material_cost_target * 100, 2) if sheet.material_cost_target else 0),
            ("人工成本", "人工工时合计", sheet.labor_cost_target, sheet.labor_cost_actual,
             sheet.labor_cost_actual - sheet.labor_cost_target,
             round((sheet.labor_cost_actual - sheet.labor_cost_target) / sheet.labor_cost_target * 100, 2) if sheet.labor_cost_target else 0),
            ("制造费用", "分摊费用合计", sheet.overhead_cost_target, sheet.overhead_cost_actual,
             sheet.overhead_cost_actual - sheet.overhead_cost_target,
             round((sheet.overhead_cost_actual - sheet.overhead_cost_target) / sheet.overhead_cost_target * 100, 2) if sheet.overhead_cost_target else 0),
        ]
        for cat_name, item_name, target_val, actual_val, variance_val, variance_pct_val in summary_data:
            cat_cell = ws.cell(row=row, column=1, value=cat_name)
            cat_cell.font = Font(name="微软雅黑", bold=True)
            cat_cell.fill = cat_fill
            cat_cell.alignment = data_align_center
            cat_cell.border = thin_border
            ws.cell(row=row, column=2, value=item_name).border = thin_border
            for col_idx, val in [(3, target_val), (4, actual_val), (5, variance_val)]:
                c = ws.cell(row=row, column=col_idx, value=val)
                c.number_format = '#,##0.00'
                c.alignment = data_align_right
                c.border = thin_border
                if col_idx == 5 and val < 0:
                    c.font = red_font
            pct_cell = ws.cell(row=row, column=6, value=variance_pct_val / 100 if variance_pct_val else 0)
            pct_cell.number_format = '0.00%'
            pct_cell.alignment = data_align_right
            pct_cell.border = thin_border
            if variance_pct_val < 0:
                pct_cell.font = red_font
            row += 1

        # 明细项
        if items_map.get(sheet.id):
            row += 1
            ws.cell(row=row, column=1, value="明细项").font = Font(name="微软雅黑", bold=True, size=10)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            row += 1
            detail_headers = ["成本类别", "明细项名称", "目标金额(元)", "实际金额(元)", "差异(元)", "差异率(%)"]
            for col_idx, h in enumerate(detail_headers, 1):
                cell = ws.cell(row=row, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
                cell.alignment = header_align
                cell.border = thin_border
            row += 1
            for item in items_map[sheet.id]:
                cat_label = categories_cn.get(item.cost_category, item.cost_category)
                ws.cell(row=row, column=1, value=cat_label).border = thin_border
                ws.cell(row=row, column=2, value=item.item_name).border = thin_border
                for col_idx, val in [(3, item.target_amount), (4, item.actual_amount), (5, item.variance)]:
                    c = ws.cell(row=row, column=col_idx, value=val or 0)
                    c.number_format = '#,##0.00'
                    c.alignment = data_align_right
                    c.border = thin_border
                    if col_idx == 5 and (val or 0) < 0:
                        c.font = red_font
                pct_val = (item.variance_pct or 0) / 100
                pct_cell = ws.cell(row=row, column=6, value=pct_val)
                pct_cell.number_format = '0.00%'
                pct_cell.alignment = data_align_right
                pct_cell.border = thin_border
                if (item.variance_pct or 0) < 0:
                    pct_cell.font = red_font
                row += 1

        # 合计行
        row += 1
        total_cells = [ws.cell(row=row, column=1, value="合计"), ws.cell(row=row, column=2, value="")]
        total_cells[0].font = total_font
        total_cells[0].fill = total_fill
        total_cells[0].alignment = data_align_center
        total_cells[0].border = thin_border
        ws.cell(row=row, column=2).border = thin_border
        for col_idx, val in [(3, sheet.total_cost_target), (4, sheet.total_cost_actual), (5, sheet.variance_amount)]:
            c = ws.cell(row=row, column=col_idx, value=val or 0)
            c.font = total_font
            c.fill = total_fill
            c.number_format = '#,##0.00'
            c.alignment = data_align_right
            c.border = thin_border
            if col_idx == 5 and (val or 0) < 0:
                c.font = Font(name="微软雅黑", bold=True, color="FF0000")
        pct_final = (sheet.variance_pct or 0) / 100
        pct_c = ws.cell(row=row, column=6, value=pct_final)
        pct_c.font = total_font
        pct_c.fill = total_fill
        pct_c.number_format = '0.00%'
        pct_c.alignment = data_align_right
        pct_c.border = thin_border
        if (sheet.variance_pct or 0) < 0:
            pct_c.font = Font(name="微软雅黑", bold=True, color="FF0000")

        # 列宽
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 14

        ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.orientation = "landscape"

    # ── 4. 输出 ──
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"成本核算_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
