"""Competitor Import/Export API — Excel/CSV 导入导出

Endpoints:
- POST  /api/pm/competitors/import   — 上传 Excel/CSV 并批量导入
- GET   /api/pm/competitors/template — 下载导入模板
- GET   /api/pm/competitors/export   — 按市场导出为 Excel
"""
import io
import csv
import logging
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    openpyxl = None

from app.core.database import get_db
from app.core.security import get_current_user
from app.core.permissions import require_role
from app.models.user import User
from app.models.competitor import CompetitorModel

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/pm/competitors", tags=["竞品库-导入导出"])

# ── 模板列定义 ──────────────────────────────────────────────────────

TEMPLATE_COLUMNS: list[dict] = [
    {"key": "brand",              "label": "品牌 *",       "width": 16},
    {"key": "model",              "label": "型号 *",       "width": 20},
    {"key": "market",             "label": "目标市场 *",   "width": 16},
    {"key": "product_type",       "label": "产品类型",     "width": 14},
    {"key": "cooling_capacity",   "label": "冷量段",       "width": 14},
    {"key": "cooling_capacity_w", "label": "制冷量(W)",    "width": 14},
    {"key": "heating_capacity_w", "label": "制热量(W)",    "width": 14},
    {"key": "energy_rating",      "label": "能效等级",     "width": 12},
    {"key": "cooling_w",          "label": "制冷功率(W)",  "width": 14},
    {"key": "heating_w",          "label": "制热功率(W)",  "width": 14},
    {"key": "eer",                "label": "EER",          "width": 10},
    {"key": "cspf",               "label": "CSPF",         "width": 10},
    {"key": "noise_indoor_db",    "label": "室内噪音(dB)", "width": 14},
    {"key": "noise_outdoor_db",   "label": "室外噪音(dB)", "width": 14},
    {"key": "airflow_m3h",        "label": "循环风量(m³/h)", "width": 16},
    {"key": "indoor_size_mm",     "label": "内机尺寸(mm)", "width": 14},
    {"key": "outdoor_size_mm",    "label": "外机尺寸(mm)", "width": 14},
    {"key": "factory_price",      "label": "出厂价",       "width": 12},
    {"key": "launch_year",        "label": "上市年份",     "width": 10},
    {"key": "notes",              "label": "备注",         "width": 20},
]

REQUIRED_KEYS = ["brand", "model", "market"]

FIELD_TYPES: dict[str, type] = {
    "cooling_capacity_w": int,
    "heating_capacity_w": int,
    "cooling_w": int,
    "heating_w": int,
    "eer": float,
    "cspf": float,
    "noise_indoor_db": float,
    "noise_outdoor_db": float,
    "airflow_m3h": float,
    "launch_year": int,
}


# ── 辅助函数 ────────────────────────────────────────────────────────

def _build_template_wb() -> openpyxl.Workbook:
    """生成模板 Excel 工作簿"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "竞品导入模板"
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ci, col in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=ci, value=col["label"])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = col["width"]

    # 第二行写示例数据
    example = [
        "TCL", "TAC-12CSF", "越南", "分体壁挂式", "12000BTU",
        3500, 4000, "5星", 900, 1000, 3.5, None,
        38, 52, 600, "800×300×200", "800×600×300", "$112", 2025, "",
    ]
    for ci, val in enumerate(example, start=1):
        ws.cell(row=2, column=ci, value=val)

    return wb


def _row_to_model(row: dict, line_num: int) -> CompetitorModel:
    """将 CSV/Excel 行数据转换为 CompetitorModel 实例"""
    item = CompetitorModel()
    for col in TEMPLATE_COLUMNS:
        key = col["key"]
        raw = row.get(key, "")
        if raw is None or raw == "":
            continue
        if key in FIELD_TYPES:
            try:
                setattr(item, key, FIELD_TYPES[key](raw))
            except (ValueError, TypeError):
                logger.warning("行 %d 字段 %s 解析失败: %s", line_num, key, raw)
        else:
            setattr(item, key, str(raw).strip())
    return item


def _validate_row(row: dict, line_num: int) -> list[str]:
    """校验单行必填字段，返回错误列表"""
    errors: list[str] = []
    for key in REQUIRED_KEYS:
        if not row.get(key) or str(row.get(key, "")).strip() == "":
            errors.append(f"行 {line_num}: 必填字段 '{key}' 为空")
    return errors


def _serialize(item: CompetitorModel) -> dict:
    """序列化单条竞品记录"""
    return {
        "id": item.id,
        "brand": item.brand,
        "model": item.model,
        "market": item.market,
        "product_type": item.product_type,
        "cooling_capacity": item.cooling_capacity,
        "cooling_capacity_w": item.cooling_capacity_w,
        "heating_capacity_w": item.heating_capacity_w,
        "energy_rating": item.energy_rating,
        "cooling_w": item.cooling_w,
        "heating_w": item.heating_w,
        "eer": item.eer,
        "cspf": item.cspf,
        "noise_indoor_db": item.noise_indoor_db,
        "noise_outdoor_db": item.noise_outdoor_db,
        "airflow_m3h": item.airflow_m3h,
        "indoor_size_mm": item.indoor_size_mm,
        "outdoor_size_mm": item.outdoor_size_mm,
        "factory_price": item.factory_price,
        "launch_year": item.launch_year,
        "notes": item.notes,
    }


def _read_excel(file_bytes: bytes) -> list[dict]:
    """解析 Excel (xlsx) 文件内容为 dict 列表"""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    if not ws:
        raise HTTPException(status_code=400, detail="Excel 文件无有效工作表")
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=400, detail="Excel 文件为空")
    headers = [str(h).strip() if h else "" for h in header_row]

    # 匹配列名到 key
    col_index: dict[int, str] = {}
    for ci, label in enumerate(headers):
        for col_def in TEMPLATE_COLUMNS:
            if col_def["label"].replace(" *", "") == label.replace(" *", ""):
                col_index[ci] = col_def["key"]
                break

    if not col_index:
        raise HTTPException(status_code=400, detail="无法匹配表头与模板列，请使用模板文件")

    results: list[dict] = []
    for row_idx, row in enumerate(rows_iter, start=2):
        record: dict = {}
        has_data = False
        for ci, val in enumerate(row):
            if ci in col_index and val is not None:
                record[col_index[ci]] = val
                if val != "":
                    has_data = True
        if has_data:
            results.append(record)

    return results


def _read_csv(file_bytes: bytes) -> list[dict]:
    """解析 CSV 内容为 dict 列表"""
    content = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content))

    # 匹配列名
    field_map: dict[str, str] = {}
    for field in reader.fieldnames or []:
        f_clean = field.strip().replace(" *", "")
        for col_def in TEMPLATE_COLUMNS:
            if col_def["label"].replace(" *", "") == f_clean:
                field_map[field] = col_def["key"]
                break

    results: list[dict] = []
    for row_idx, row in enumerate(reader, start=2):
        record: dict = {}
        has_data = False
        for src_field, val in row.items():
            mapped = field_map.get(src_field)
            if mapped and val and val.strip():
                record[mapped] = val.strip()
                has_data = True
        if has_data:
            results.append(record)

    return results


# ── API 端点 ────────────────────────────────────────────────────────


@router.get("/template")
def download_template(
    current_user: User = Depends(get_current_user),
    _=Depends(require_role("admin", "product_manager")),
) -> StreamingResponse:
    """下载竞品导入模板 (Excel)"""
    wb = _build_template_wb()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=competitor_import_template.xlsx"},
    )


@router.post("/import")
def import_competitors(
    file: UploadFile = File(..., description="Excel(.xlsx) 或 CSV(.csv) 文件"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    _=Depends(require_role("admin", "product_manager")),
) -> dict:
    """批量导入竞品数据（支持 xlsx / csv）"""
    file_bytes = file.file.read()

    # 检测文件类型
    ext = (file.filename or "").lower()
    if ext.endswith(".csv"):
        records = _read_csv(file_bytes)
    elif ext.endswith(".xlsx"):
        records = _read_excel(file_bytes)
    else:
        raise HTTPException(status_code=400, detail=f"不支持的文件格式: {ext}，请使用 .xlsx 或 .csv")

    if not records:
        raise HTTPException(status_code=400, detail="文件中无有效数据行")

    # 校验并导入
    errors: list[str] = []
    created: list[CompetitorModel] = []
    for idx, record in enumerate(records, start=2):
        row_errors = _validate_row(record, idx)
        if row_errors:
            errors.extend(row_errors)
            continue
        try:
            item = _row_to_model(record, idx)
            db.add(item)
            db.flush()
            created.append(item)
        except Exception as e:
            logger.exception("unexpected error")
            errors.append(f"行 {idx}: 导入失败 — {str(e)}")

    if created:
        db.commit()
        # 刷新以获取 ID
        for item in created:
            db.refresh(item)

    result = {
        "total": len(records),
        "imported": len(created),
        "errors": len(errors),
        "error_details": errors[:20],  # 最多返回20条错误
        "items": [_serialize(it) for it in created],
    }
    return result


@router.get("/export")
def export_competitors(
    market: Optional[str] = Query(None, description="目标市场过滤"),
    brand: Optional[str] = Query(None, description="品牌过滤"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> StreamingResponse:
    """导出竞品数据为 Excel（支持按市场/品牌过滤）"""
    q = db.query(CompetitorModel)
    if market:
        q = q.filter(CompetitorModel.market == market)
    if brand:
        q = q.filter(CompetitorModel.brand == brand)
    items = q.order_by(CompetitorModel.id.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "竞品数据"

    # 表头
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for ci, col_def in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=ci, value=col_def["label"])
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = col_def["width"]

    # 数据行
    for ri, item in enumerate(items, start=2):
        for ci, col_def in enumerate(TEMPLATE_COLUMNS, start=1):
            val = getattr(item, col_def["key"], None)
            if val is not None:
                ws.cell(row=ri, column=ci, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename_parts = ["competitors"]
    if market:
        filename_parts.append(market)
    filename = "_".join(filename_parts) + ".xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
