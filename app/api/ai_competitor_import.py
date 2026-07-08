"""竞品 AI 导入 API — 文件上传 → AI 提取 → 预览

基于 LLMParamExtractor 的智能提取流程，支持文件上传、AI 解析预览。

Endpoints:
- POST /competitors/ai-import/upload         — 上传文件(excel/pdf/csv)
- POST /competitors/ai-import/{session_id}/parse  — 触发 AI 提取
- GET  /competitors/ai-import/{session_id}/preview — 获取预览结果
"""
import io
import re
import uuid
import logging
from datetime import datetime, timezone
from typing import Any, Optional

# ── 中文参数名 → 英文参数名映射（用于 Excel 矩阵表格）──
CN_PARAM_MAP: dict[str, str] = {
    "系列名": "Series",
    "型号编码": "Model Code",
    "制冷量T1 (Btu/h)": "Cooling Capacity T1 (BTU/h)",
    "制冷量T1 (W)": "Cooling Capacity T1 (W)",
    "制冷量T3 (Btu/h)": "Cooling Capacity T3 (BTU/h)",
    "制冷量T3 (W)": "Cooling Capacity T3 (W)",
    "最小制冷量(Btu/h)": "Min Cooling (BTU/h)",
    "最大制冷量(Btu/h)": "Max Cooling (BTU/h)",
    "制热量(W)": "Heating Capacity (W)",
    "T1功率输入(W)": "Power Input T1 (W)",
    "T1电流(A)": "Current T1 (A)",
    "T3功率输入(W)": "Power Input T3 (W)",
    "T3电流(A)": "Current T3 (A)",
    "室内风量(m³/h)": "Indoor Airflow (m3/h)",
    "室内机尺寸(mm)": "Indoor Unit Size (mm)",
    "室内机重量(kg)": "Indoor Unit Weight (kg)",
    "噪音dB(A)": "Indoor Noise (dB)",
    "SEER": "SEER",
    "电源": "Power Supply",
    "室外机尺寸(mm)": "Outdoor Unit Size (mm)",
    "室外机重量(kg)": "Outdoor Unit Weight (kg)",
    "制冷剂": "Refrigerant",
    "压缩机": "Compressor",
    "液管(Inch)": "Liquid Pipe (inch)",
    "气管(Inch)": "Gas Pipe (inch)",
    "装柜量(40HC)": "Container Qty (40HC)",
    "网址": "URL",
}

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.security import get_current_user, require_role
from app.models.user import User
from app.models.target_market import TargetMarket
from app.services.competitor_ai_parser import CompetitorAIParser
from app.schemas.ai_competitor_import import (
    ImportSourceType,
    AICompetitorImportResponse,
)

logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/pm/ai-import",
    tags=["竞品AI导入"],
)

# ── 内存会话存储（生产环境应替换为 Redis/DB） ──
# session_id -> {
#   "source_type": ImportSourceType,
#   "target_market_id": int,
#   "market_code": str,
#   "original_filename": str | None,
#   "raw_text": str,
#   "extractions": list | None,
#   "created_at": datetime,
#   "status": "parsing" | "preview" | "completed" | "failed",
# }
_sessions: dict[str, dict[str, Any]] = {}

# ── 支持的解析文件扩展 ──
TEXT_EXTENSIONS = {".txt", ".csv", ".json", ".md", ".html", ".htm"}
PDF_EXTENSION = {".pdf"}
EXCEL_EXTENSIONS = {".xlsx", ".xls"}
ALLOWED_EXTENSIONS = TEXT_EXTENSIONS | PDF_EXTENSION | EXCEL_EXTENSIONS

MAX_FILE_SIZE = 20 * 1024 * 1024  # 20 MB


# ═══════════════════════════════════════════════════════════════════════
# 辅助函数
# ═══════════════════════════════════════════════════════════════════════


def _resolve_market_code(db: Session, target_market_id: int) -> str:
    """根据 target_market_id 查询 market_code

    Args:
        db: 数据库 Session
        target_market_id: 目标市场主键

    Returns:
        str: 市场代码

    Raises:
        HTTPException: 市场不存在
    """
    market = db.query(TargetMarket).filter(
        TargetMarket.id == target_market_id
    ).first()
    if not market:
        raise HTTPException(
            status_code=404,
            detail=f"目标市场不存在 (id={target_market_id})",
        )
    return market.market_code


def _resolve_market_by_name_or_code(db: Session, market: str) -> int:
    """根据市场名称或代码查询 target_market_id

    Args:
        db: 数据库 Session
        market: 市场名称（如"越南"）或代码（如"VN"）

    Returns:
        int: target_market_id

    Raises:
        HTTPException: 市场不存在
    """
    tm = db.query(TargetMarket).filter(
        (TargetMarket.market_name == market) | (TargetMarket.market_code == market)
    ).first()
    if not tm:
        raise HTTPException(
            status_code=404,
            detail=f"目标市场不存在 (market={market})",
        )
    return tm.id


def _extract_text_from_file(
    file_bytes: bytes,
    filename: str,
) -> str:
    """从上传文件提取纯文本

    支持 txt/csv/pdf/xlsx/md/html，自动处理中文编码、PDF与Excel矩阵格式。

    Raises:
        HTTPException: 不支持的文件类型
    """
    ext = (filename or "").lower()
    # 提取扩展名
    for known_ext in sorted(ALLOWED_EXTENSIONS, key=len, reverse=True):
        if ext.endswith(known_ext):
            ext = known_ext
            break
    else:
        # 尝试取最后一个 . 后的部分
        dot_idx = ext.rfind(".")
        if dot_idx != -1:
            ext = ext[dot_idx:]
        else:
            raise HTTPException(
                status_code=400,
                detail=f"不支持的文件格式，请使用 {', '.join(sorted(ALLOWED_EXTENSIONS))}",
            )

    # ── 文本类文件 ──
    if ext in TEXT_EXTENSIONS:
        try:
            return file_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            # 尝试 gbk 编码（中文环境）
            try:
                return file_bytes.decode("gbk")
            except UnicodeDecodeError:
                return file_bytes.decode("utf-8", errors="replace")

    # ── PDF ──
    if ext in PDF_EXTENSION:
        try:
            import fitz  # pymupdf — 远超 PyPDF2 的文本/表格提取能力
        except ImportError:
            raise HTTPException(
                status_code=400,
                detail="PDF 解析需要安装 pymupdf: pip install pymupdf",
            )
        text_parts: list[str] = []
        try:
            doc = fitz.open(stream=file_bytes, filetype="pdf")
            for page in doc:
                # pymupdf 两阶段提取：文本层 + 图像 OCR 文本块
                page_text = page.get_text("text")
                if page_text and page_text.strip():
                    text_parts.append(page_text)
                else:
                    # 空白页／纯图片页尝试 OCR
                    try:
                        ocr_text = page.get_text("dict")
                        blocks = [b for b in ocr_text.get("blocks", []) if b.get("type") == 0]
                        if blocks:
                            lines = []
                            for b in blocks:
                                for line in b.get("lines", []):
                                    spans = [s.get("text", "") for s in line.get("spans", [])]
                                    lines.append(" ".join(spans))
                            text_parts.append("\n".join(lines))
                    except Exception:
                        logger.debug(f"ignored: {{e}}")
                        pass
            doc.close()
        except Exception as exc:
            logger.warning("PDF 解析失败: %s", exc)
            raise HTTPException(
                status_code=400, detail=f"PDF 文件解析失败: {exc}",
            )
        return "\n".join(text_parts)

    # ── Excel ──
    if ext in EXCEL_EXTENSIONS:
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(
                status_code=400,
                detail="Excel 解析需要安装 openpyxl: pip install openpyxl",
            )
        text_parts = []
        try:
            wb = openpyxl.load_workbook(
                io.BytesIO(file_bytes), read_only=True, data_only=True,
            )
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # 读取全部单元格为 2D 数组
                all_rows: list[list[str]] = []
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    all_rows.append(cells)

                # ── 检测矩阵格式（列式表格：第一数据行是序号 1,2,3...） ──
                # 跳过空行，找到第一个有内容的行
                data_start = 0
                while data_start < len(all_rows) and all(
                    not c.strip() for c in all_rows[data_start]
                ):
                    data_start += 1

                is_matrix = False
                if data_start < len(all_rows) and len(all_rows[data_start]) > 3:
                    first_data_row = all_rows[data_start]
                    # 跳过前 2 列（空列 + "序"列），从第 3 列检查数字序号
                    numeric_cols = 0
                    total_check = 0
                    for c in first_data_row[2:]:  # 从第 3 列起
                        total_check += 1
                        try:
                            int(c.strip())
                            numeric_cols += 1
                        except ValueError:
                            pass  # 不 break，继续检查
                    # 如果超过 60% 的列是数字序号 → 矩阵格式
                    if total_check >= 3 and numeric_cols >= total_check * 0.6:
                        is_matrix = True

                if is_matrix:
                    # ══════════════════════════════════════════════════
                    # 矩阵格式 → 每列是一个产品，转换为产品描述文本
                    # ══════════════════════════════════════════════════
                    num_cols = len(all_rows[data_start])
                    # 收集所有有数据的行（跳过空行）
                    param_rows: list[list[str]] = []
                    for r in all_rows[data_start:]:
                        # 跳过全空行
                        if all(not c.strip() for c in r):
                            continue
                        # 填充到相同长度
                        while len(r) < num_cols:
                            r.append("")
                        param_rows.append(r)

                    if len(param_rows) >= 3:
                        # 第0行：序号行（跳过）
                        # 第1行：系列名行（填充 forward-fill 已合并单元格的值）
                        series_row = param_rows[1] if len(param_rows) > 1 else []
                        # 第2行：型号编码行
                        model_row = param_rows[2] if len(param_rows) > 2 else []

                        # Forward-fill 系列名（处理合并单元格）
                        last_series = ""
                        for ci in range(len(series_row)):
                            v = series_row[ci].strip()
                            if v:
                                last_series = v
                            series_row[ci] = last_series

                        # 对每个产品列（从列 2 开始）构建描述文本，col 0=空 col 1=参数名
                        for ci in range(2, num_cols):
                            product_index = ci - 1  # 从 1 开始编号
                            series = series_row[ci] if ci < len(series_row) else ""
                            model = model_row[ci] if ci < len(model_row) else ""
                            lines = [f"=== Product {product_index}: {series} {model} ==="]
                            if series:
                                lines.append(f"Series: {series}")
                            if model:
                                lines.append(f"Model: {model}")

                            # 从第 3 行起是参数行，参数名在 col 1，参数值在 col ci
                            for ri in range(3, len(param_rows)):
                                cn_name = param_rows[ri][1].strip() if len(param_rows[ri]) > 1 else ""
                                param_name = CN_PARAM_MAP.get(cn_name, cn_name)
                                param_val = param_rows[ri][ci].strip() if ci < len(param_rows[ri]) else ""
                                if param_name and param_val:
                                    lines.append(f"{param_name}: {param_val}")

                            text_parts.append("\n".join(lines))
                    else:
                        # 行数不足3行，用原始格式
                        text_parts.append(
                            f"=== Sheet: {sheet_name} ===\n"
                            + "\n".join("\t".join(r) for r in all_rows)
                        )
                else:
                    # ══════════════════════════════════════════════════
                    # 非矩阵格式 → 逐行 tab 分隔（原始行为）
                    # ══════════════════════════════════════════════════
                    text_parts.append(
                        f"=== Sheet: {sheet_name} ===\n"
                        + "\n".join("\t".join(r) for r in all_rows)
                    )
        except Exception as exc:
            logger.warning("Excel 解析失败: %s", exc)
            raise HTTPException(
                status_code=400, detail=f"Excel 文件解析失败: {exc}",
            )
        return "\n".join(text_parts)

    # 不应到达这里
    raise HTTPException(
        status_code=400,
        detail=f"不支持的文件格式: {ext}",
    )


def _split_product_sections(text: str) -> list[str]:
    """将包含多个产品的文本拆分为产品段落列表

    策略：按 Model:/Modello/Product 等常见产品表头拆分，过短段落被过滤。
    """
    if not text or len(text) < 100:
        return [text] if text else []

    # 尝试按产品表头拆分（匹配多语言的产品/型号标签）
    # 常见模式: "Modello/Model CLIMADESIGN 12", "Model: XYZ-123", "型号:"
    split_pattern = re.compile(
        r'(?:'
        r'Modello/Model\s+\w+[\s\d]+|'       # 意大利/英文 "Modello/Model CLIMADESIGN 12"
        r'Model\s*(?::|Number)?\s*\w+[\s\d]+|'  # 英文 "Model: ABC-123"
        r'型号[：:\s]+\w+|'                    # 中文 "型号：KFR-35GW"
        r'Product\s*(?::|Name)?\s*\w+'         # 英文 "Product: XYZ"
        r')',
        re.IGNORECASE,
    )

    matches = list(split_pattern.finditer(text))
    if len(matches) <= 1:
        # 无法拆分，返回整个文本
        return [text]

    sections: list[str] = []
    prev_start = matches[0].start()
    for i, m in enumerate(matches):
        if i == 0:
            continue
        start = m.start()
        # 前一段从上一个匹配位置到本次匹配位置
        section_text = text[prev_start:start].strip()
        if len(section_text) >= 50:
            sections.append(section_text)
        prev_start = start

    # 最后一段
    last_section = text[prev_start:].strip()
    if len(last_section) >= 50:
        sections.append(last_section)

    # 如果拆分后只有一个段落，退回到不拆分
    if len(sections) <= 1:
        return [text]

    logger.info("文本拆分为 %d 个产品段落", len(sections))
    return sections


# ═══════════════════════════════════════════════════════════════════════
# API 端点
# ═══════════════════════════════════════════════════════════════════════


@router.post("/upload", response_model=dict)
async def upload_file(
    file: UploadFile = File(..., description="竞品文件（txt/csv/pdf/xlsx/md/html）"),
    market: str = Query(..., description="目标市场代码或名称"),
    source_type: ImportSourceType = Query(
        ImportSourceType.FILE, description="导入源类型",
    ),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    _=Depends(require_role("admin", "product_manager")),
) -> dict:
    """上传竞品文件并创建解析会话

    支持 txt / csv / pdf / xlsx / md / html 格式，自动提取文本内容。
    返回 session_id 供后续解析调用。
    """
    # ── 解析市场 ID ──
    target_market_id = _resolve_market_by_name_or_code(db, market)

    # ── 校验文件 ──
    if not file.filename:
        raise HTTPException(status_code=400, detail="文件名不能为空")

    ext = (file.filename or "").lower()
    if not any(ext.endswith(e) for e in ALLOWED_EXTENSIONS):
        raise HTTPException(
            status_code=400,
            detail=f"不支持的文件格式，支持: {', '.join(sorted(ALLOWED_EXTENSIONS))}",
        )

    file_bytes = await file.read()
    if len(file_bytes) > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=400,
            detail=f"文件过大，最大支持 {MAX_FILE_SIZE // 1024 // 1024}MB",
        )

    # ── 提取文本 ──
    try:
        raw_text = _extract_text_from_file(file_bytes, file.filename)
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("文件文本提取失败: %s", exc, exc_info=True)
        raise HTTPException(
            status_code=400, detail=f"文件文本提取失败: {exc}",
        )

    if not raw_text.strip():
        raise HTTPException(
            status_code=400, detail="文件中未提取到有效文本内容",
        )

    # ── 创建会话 ──
    session_id = uuid.uuid4().hex
    _sessions[session_id] = {
        "source_type": source_type,
        "target_market_id": target_market_id,
        "original_filename": file.filename,
        "raw_text": raw_text,
        "extractions": None,
        "created_at": datetime.now(timezone.utc),
        "status": "parsing",
        "total_imported": 0,
        "total_skipped": 0,
    }

    logger.info(
        "AI导入会话创建: session_id=%s, file=%s, market_id=%d, text_len=%d",
        session_id, file.filename, target_market_id, len(raw_text),
    )

    return {
        "session_id": session_id,
        "target_market_id": target_market_id,
        "filename": file.filename,
        "text_length": len(raw_text),
        "text_snippet": raw_text[:300] if raw_text else "",
        "message": "文件上传成功，请调用 /parse 触发 AI 提取",
    }


@router.post("/{session_id}/parse", response_model=AICompetitorImportResponse)
async def trigger_parse(
    session_id: str,
    brand: Optional[str] = Query(None, description="品牌名称（兜底）"),
    provider: Optional[str] = Query(None, description="AI 供应商"),
    model: Optional[str] = Query(None, description="AI 模型名"),
    api_key: Optional[str] = Query(None, description="API 密钥"),
    api_base: Optional[str] = Query(None, description="API Base URL"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    _=Depends(require_role("admin", "product_manager")),
) -> AICompetitorImportResponse:
    """触发 AI 提取 — 对上传的文本执行 LLM 参数解析

    使用 LLMParamExtractor 提取空调参数，返回结构化预览结果。
    可选参数：brand（品牌兜底）、provider/model/api_key/api_base（AI 配置）。
    """
    # ── 校验会话 ──
    session = _sessions.get(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="会话不存在或已过期")

    if session["status"] != "parsing":
        raise HTTPException(
            status_code=400,
            detail=f"会话状态不允许解析 (当前状态: {session['status']})",
        )

    raw_text: str = session["raw_text"]
    target_market_id: int = session["target_market_id"]
    source_type: ImportSourceType = session["source_type"]

    # ── 获取市场代码 ──
    market_code = _resolve_market_code(db, target_market_id)

    # ── 调用 AI 解析（单次调用，支持多产品提取） ──
    parser = CompetitorAIParser(db=db)
    try:
        # 使用多产品提取 — 一次 LLM 调用返回所有产品
        extractions = await parser.parse_text_multi(
            raw_text=raw_text,
            market_code=market_code,
            default_brand=brand,
            provider=provider,
            model=model,
            api_key=api_key,
            api_base=api_base,
        )

        if not extractions:
            extractions = []
            logger.warning("AI 提取未返回任何结果")

        session["status"] = "preview"
        session["extractions"] = extractions
    except Exception as exc:
        logger.error("AI 提取失败: %s", exc, exc_info=True)
        session["status"] = "failed"
        raise HTTPException(
            status_code=500,
            detail=f"AI 提取失败: {exc}",
        )

    logger.info(
        "AI提取完成: session_id=%s, extractions=%d, first_confidence=%.2f",
        session_id, len(extractions),
        extractions[0].overall_confidence if extractions else 0,
    )

    return AICompetitorImportResponse(
        session_id=session_id,
        source_type=source_type,
        target_market_id=target_market_id,
        total_extracted=len(extractions),
        extractions=extractions,
    )


@router.get("/{session_id}/preview", response_model=AICompetitorImportResponse)
def preview_extractions(
    session_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    _=Depends(require_role("admin", "product_manager")),
) -> AICompetitorImportResponse:
    """获取 AI 提取预览结果

    在确认导入前查看提取到的竞品参数明细。
    """
    session = _sessions.get(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="会话不存在或已过期")

    if session["status"] not in ("preview", "completed"):
        raise HTTPException(
            status_code=400,
            detail=f"会话尚未完成解析 (当前状态: {session['status']})",
        )

    extractions = session.get("extractions") or []
    return AICompetitorImportResponse(
        session_id=session_id,
        source_type=session["source_type"],
        target_market_id=session["target_market_id"],
        total_extracted=len(extractions),
        extractions=extractions,
    )
